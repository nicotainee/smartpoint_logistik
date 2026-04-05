from flask import Flask, render_template, request, redirect, url_for, flash, session
import mysql.connector
import qrcode
import os
from werkzeug.utils import secure_filename
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'smartpoint_secret_key_2024')

UPLOAD_FOLDER = os.path.join('static', 'foto_rumah')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ─── KREDENSIAL ADMIN ──────────────────────────────────────────────────────
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'smartpoint2024')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    return mysql.connector.connect(
        host=os.environ.get('MYSQLHOST', 'localhost'),
        user=os.environ.get('MYSQLUSER', 'root'),
        password=os.environ.get('MYSQLPASSWORD', ''),
        database=os.environ.get('MYSQLDATABASE', 'railway'),
        port=int(os.environ.get('MYSQLPORT', 3306))
    )

# ─── DECORATOR LOGIN REQUIRED ──────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ─── LOGIN / LOGOUT ────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('logged_in'):
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Username atau password salah!')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── LANDING PAGE ─────────────────────────────────────────────────────────

@app.route('/')
def landing():
    return render_template('landing.html')


# ─── HALAMAN KURIR (QR Scan) ───────────────────────────────────────────────

@app.route('/cek/<kode_unik>')
def cek_alamat(kode_unik):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM smart_points WHERE kode_unik = %s", (kode_unik,))
    data = cursor.fetchone()
    conn.close()
    if data:
        return render_template('index.html', alamat=data)
    return render_template('404.html'), 404


# ─── DASHBOARD ADMIN ───────────────────────────────────────────────────────

@app.route('/admin')
@login_required
def admin_dashboard():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM smart_points ORDER BY id DESC")
    data = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) as total FROM smart_points")
    total = cursor.fetchone()['total']
    conn.close()
    return render_template('admin.html', smart_points=data, total=total)


@app.route('/admin/tambah', methods=['GET', 'POST'])
@login_required
def tambah_alamat():
    if request.method == 'POST':
        kode    = request.form['kode_unik']
        nama    = request.form['nama_penerima']
        lat     = request.form['latitude']
        lon     = request.form['longitude']
        patokan = request.form['patokan_visual']

        nama_file_foto = None
        if 'foto_rumah' in request.files:
            foto = request.files['foto_rumah']
            if foto and foto.filename != '' and allowed_file(foto.filename):
                ext = foto.filename.rsplit('.', 1)[1].lower()
                nama_file_foto = f"foto_{kode}.{ext}"
                foto.save(os.path.join(UPLOAD_FOLDER, nama_file_foto))

        conn = get_db()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO smart_points (kode_unik, nama_penerima, latitude, longitude, patokan_visual, link_foto_rumah) VALUES (%s,%s,%s,%s,%s,%s)",
                (kode, nama, lat, lon, patokan, nama_file_foto)
            )
            conn.commit()
            maps_url = f"https://www.google.com/maps?q={lat},{lon}"
            img = qrcode.make(maps_url)
            img.save(os.path.join('static', f'QR_{kode}.png'))
            flash(f'Smart Point "{kode}" berhasil ditambahkan!', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'error')
        finally:
            conn.close()
        return redirect(url_for('admin_dashboard'))

    return render_template('tambah.html')


@app.route('/admin/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_alamat(id):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        kode    = request.form['kode_unik']
        nama    = request.form['nama_penerima']
        lat     = request.form['latitude']
        lon     = request.form['longitude']
        patokan = request.form['patokan_visual']

        nama_file_foto = request.form.get('foto_lama')
        if 'foto_rumah' in request.files:
            foto = request.files['foto_rumah']
            if foto and foto.filename != '' and allowed_file(foto.filename):
                ext = foto.filename.rsplit('.', 1)[1].lower()
                nama_file_foto = f"foto_{kode}.{ext}"
                foto.save(os.path.join(UPLOAD_FOLDER, nama_file_foto))

        try:
            cursor.execute(
                "UPDATE smart_points SET kode_unik=%s, nama_penerima=%s, latitude=%s, longitude=%s, patokan_visual=%s, link_foto_rumah=%s WHERE id=%s",
                (kode, nama, lat, lon, patokan, nama_file_foto, id)
            )
            conn.commit()
            flash(f'Data "{kode}" berhasil diupdate!', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'error')
        finally:
            conn.close()
        return redirect(url_for('admin_dashboard'))

    cursor.execute("SELECT * FROM smart_points WHERE id = %s", (id,))
    data = cursor.fetchone()
    conn.close()
    if not data:
        flash('Data tidak ditemukan.', 'error')
        return redirect(url_for('admin_dashboard'))
    return render_template('edit.html', point=data)


@app.route('/admin/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_alamat(id):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM smart_points WHERE id = %s", (id,))
        conn.commit()
        flash('Data berhasil dihapus.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/qr/<kode_unik>')
@login_required
def lihat_qr(kode_unik):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM smart_points WHERE kode_unik = %s", (kode_unik,))
    data = cursor.fetchone()
    conn.close()
    if not data:
        flash('Data tidak ditemukan.', 'error')
        return redirect(url_for('admin_dashboard'))

    qr_file = f'QR_{kode_unik}.png'
    qr_path = os.path.join('static', qr_file)
    if not os.path.exists(qr_path):
        maps_url = f"https://www.google.com/maps?q={data['latitude']},{data['longitude']}"
        img = qrcode.make(maps_url)
        img.save(qr_path)

    return render_template('qr.html', point=data, qr_file=qr_file)


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))


# ─── UPLOAD EXCEL MASSAL ───────────────────────────────────────────────────

@app.route('/admin/upload-excel', methods=['GET', 'POST'])
@login_required
def upload_excel():
    hasil = None
    if request.method == 'POST':
        if 'file_excel' not in request.files:
            flash('Pilih file Excel dulu!', 'error')
            return redirect(request.url)

        file = request.files['file_excel']
        if file.filename == '':
            flash('Pilih file Excel dulu!', 'error')
            return redirect(request.url)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
            required = ['kode_unik', 'nama_penerima', 'latitude', 'longitude']
            for r in required:
                if r not in headers:
                    flash(f'Kolom "{r}" tidak ditemukan di Excel!', 'error')
                    return redirect(request.url)

            def col(row, name):
                idx = headers.index(name)
                val = row[idx].value
                return str(val).strip() if val is not None else ''

            conn = get_db()
            cursor = conn.cursor()
            berhasil = []
            gagal = []

            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if not any(cell.value for cell in row):
                    continue
                kode = col(row, 'kode_unik')
                nama = col(row, 'nama_penerima')
                lat  = col(row, 'latitude')
                lon  = col(row, 'longitude')
                patokan = col(row, 'patokan_visual') if 'patokan_visual' in headers else ''

                if not all([kode, nama, lat, lon]):
                    gagal.append(f'Baris {i}: data tidak lengkap')
                    continue

                try:
                    cursor.execute(
                        "INSERT INTO smart_points (kode_unik, nama_penerima, latitude, longitude, patokan_visual) VALUES (%s,%s,%s,%s,%s)",
                        (kode, nama, lat, lon, patokan)
                    )
                    conn.commit()

                    # Generate QR
                    maps_url = f"https://www.google.com/maps?q={lat},{lon}"
                    img = qrcode.make(maps_url)
                    img.save(os.path.join('static', f'QR_{kode}.png'))
                    berhasil.append(kode)
                except Exception as e:
                    gagal.append(f'Baris {i} ({kode}): {str(e)[:60]}')

            conn.close()
            hasil = {'berhasil': berhasil, 'gagal': gagal}

        except Exception as e:
            flash(f'Error baca file: {e}', 'error')

    return render_template('upload_excel.html', hasil=hasil)


@app.route('/admin/template-excel')
@login_required
def template_excel():
    import xlsxwriter, io
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet('SmartPoint')

    # Header style
    header_fmt = wb.add_format({'bold': True, 'bg_color': '#00e5a0', 'font_color': '#0d0f14', 'border': 1, 'align': 'center'})
    example_fmt = wb.add_format({'bg_color': '#1a1f2e', 'font_color': '#a0a8c0', 'border': 1})

    headers = ['kode_unik', 'nama_penerima', 'latitude', 'longitude', 'patokan_visual']
    widths  = [15, 25, 15, 15, 40]
    examples = [['BDG001', 'Budi Santoso', '-6.974028', '107.630367', 'Pagar biru, rumah pojok kanan'],
                ['BDG002', 'Siti Rahayu',  '-6.920000', '107.610000', 'Pagar hitam, ada pohon mangga']]

    for col_idx, (h, w) in enumerate(zip(headers, widths)):
        ws.write(0, col_idx, h, header_fmt)
        ws.set_column(col_idx, col_idx, w)

    for row_idx, row in enumerate(examples, start=1):
        for col_idx, val in enumerate(row):
            ws.write(row_idx, col_idx, val, example_fmt)

    wb.close()
    output.seek(0)

    from flask import send_file
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template_smartpoint.xlsx')
